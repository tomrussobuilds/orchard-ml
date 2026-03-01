"""
Study Result Export Functions.

Handles serialization of Optuna study results to various formats:

- Best trial configuration (YAML)
- Complete study metadata (JSON)
- Top K trials comparison (Excel)

All export functions handle edge cases (no completed trials, missing
timestamps) and provide informative logging with professional Excel formatting.
"""

from __future__ import annotations

import json
import logging
import math
from pathlib import Path
from typing import cast

import optuna
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ...core import LOGGER_NAME, Config, LogStyle, RunPaths, save_config_as_yaml
from .config import map_param_to_config_path
from .utils import get_completed_trials, has_completed_trials

logger = logging.getLogger(LOGGER_NAME)


# CONFIG EXPORT
def export_best_config(study: optuna.Study, cfg: Config, paths: RunPaths) -> Path | None:
    """
    Export best trial configuration as YAML file.

    Creates a new Config instance with best hyperparameters applied,
    validates it, and saves to reports/best_config.yaml.

    Args:
        study: Completed Optuna study with at least one successful trial
        cfg: Template configuration (used for non-optimized parameters)
        paths: RunPaths instance for output location

    Returns:
        Path to exported config file, or None if no completed trials

    Note:
        Skips export with warning if no completed trials exist.

    Example:
        >>> export_best_config(study, cfg, paths)
        # Creates: {paths.reports}/best_config.yaml
    """
    if not has_completed_trials(study):
        logger.warning("No completed trials. Cannot export best config.")
        return None

    # Build config dict with best parameters
    config_dict = build_best_config_dict(study.best_params, cfg)

    # Create and validate new config
    best_config = Config(**config_dict)

    # Save to YAML
    output_path = paths.reports / "best_config.yaml"
    save_config_as_yaml(best_config, output_path)

    return output_path


def export_study_summary(study: optuna.Study, paths: RunPaths) -> None:
    """
    Export complete study metadata to JSON.

    Serializes all trials with parameters, values, states, timestamps,
    and durations. Handles studies with zero completed trials gracefully.

    Args:
        study: Optuna study (may contain failed/pruned trials)
        paths: RunPaths instance for output location

    Output structure::

        {
            "study_name": str,
            "direction": str,
            "n_trials": int,
            "n_completed": int,
            "best_trial": {...} or null,
            "trials": [...]
        }

    Example:
        >>> export_study_summary(study, paths)
        # Creates: {paths.reports}/study_summary.json
    """
    completed = get_completed_trials(study)

    # Build best trial data (may be None if no completed trials)
    best_trial_data = build_best_trial_data(study, completed)

    summary = {
        "study_name": study.study_name,
        "direction": study.direction.name,
        "n_trials": len(study.trials),
        "n_completed": len(completed),
        "best_trial": best_trial_data,
        "trials": [build_trial_data(trial) for trial in study.trials],
    }

    output_path = paths.reports / "study_summary.json"
    with open(output_path, "w") as f:
        json.dump(summary, f, indent=2)

    logger.info(  # pragma: no mutant
        f"{LogStyle.INDENT}{LogStyle.ARROW} {'Study Summary':<22}: {Path(output_path).name}"
    )


def export_top_trials(
    study: optuna.Study, paths: RunPaths, metric_name: str, top_k: int = 10
) -> None:
    """
    Export top K trials to Excel spreadsheet with professional formatting.

    Creates human-readable comparison table of best-performing trials
    with hyperparameters, metric values, and durations. Applies professional
    Excel styling matching TrainingReport format.

    Args:
        study: Completed Optuna study with at least one successful trial
        paths: RunPaths instance for output location
        metric_name: Name of optimization metric (for column header)
        top_k: Number of top trials to export (default: 10)

    DataFrame Columns:

    - Rank: 1-based ranking
    - Trial: Trial number
    - {METRIC_NAME}: Objective value
    - {param_name}: Each hyperparameter
    - Duration (s): Trial duration if available

    Example:
        >>> export_top_trials(study, paths, "auc", top_k=10)
        # Creates: {paths.reports}/top_10_trials.xlsx
    """
    completed = get_completed_trials(study)
    if not completed:
        logger.warning("No completed trials. Cannot export top trials.")
        return

    reverse = study.direction == optuna.study.StudyDirection.MAXIMIZE
    # Filter out trials with None or NaN values before sorting
    valid_trials = [
        t
        for t in completed
        if t.value is not None and not (isinstance(t.value, float) and math.isnan(t.value))
    ]
    sorted_trials = sorted(valid_trials, key=lambda t: cast(float, t.value), reverse=reverse)[
        :top_k
    ]

    df = build_top_trials_dataframe(sorted_trials, metric_name)

    output_path = paths.reports / "top_10_trials.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Top Trials"

    _write_styled_rows(ws, df)
    _auto_adjust_column_widths(ws)

    wb.save(output_path)
    logger.info(  # pragma: no mutant
        f"{LogStyle.INDENT}{LogStyle.ARROW} {'Top Trials':<22}: {Path(output_path).name} "
        f"({len(sorted_trials)} trials)"
    )


def _write_styled_rows(ws, df: pd.DataFrame) -> None:
    """
    Write DataFrame rows to worksheet with professional formatting.

    Applies header styling (green fill, bold, centered) and body styling
    (left-aligned, number formatting for floats/ints).

    Args:
        ws: Active openpyxl worksheet.
        df: DataFrame to write.
    """
    header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alignment_center = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_left

                if isinstance(value, float):
                    cell.number_format = "0.0000"
                elif isinstance(value, int) and not isinstance(value, bool):
                    cell.number_format = "0"


def _auto_adjust_column_widths(ws) -> None:
    """
    Auto-adjust column widths based on cell content length.

    Args:
        ws: Active openpyxl worksheet with populated cells.
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)


# ==================== HELPER FUNCTIONS ====================


def build_best_config_dict(best_params: dict, cfg: Config) -> dict:
    """
    Construct config dictionary from best trial parameters.

    Maps Optuna parameters back to Config structure using
    map_param_to_config_path and restores full training epochs.

    Args:
        best_params: Dictionary from study.best_params
        cfg: Template config for structure and defaults

    Returns:
        Config dictionary ready for validation
    """
    config_dict = cfg.model_dump()

    for param_name, value in best_params.items():
        section, key = map_param_to_config_path(param_name)
        config_dict[section][key] = value

    # Restore normal epochs for final training (not Optuna short epochs)
    config_dict["training"]["epochs"] = cfg.training.epochs

    return config_dict  # type: ignore[no-any-return]


def build_best_trial_data(
    study: optuna.Study, completed: list[optuna.trial.FrozenTrial]
) -> dict | None:
    """
    Build best trial metadata dictionary.

    Args:
        study: Optuna study instance
        completed: list of completed trials

    Returns:
        Dictionary with best trial info, or None if no completed trials
    """
    if not completed:
        return None

    try:
        best = study.best_trial
        return {
            "number": best.number,
            "value": best.value,
            "params": best.params,
            "datetime_start": best.datetime_start.isoformat() if best.datetime_start else None,
            "datetime_complete": (
                best.datetime_complete.isoformat() if best.datetime_complete else None
            ),
        }
    except ValueError:
        # No best trial available
        return None


def build_trial_data(trial: optuna.trial.FrozenTrial) -> dict:
    """
    Build trial metadata dictionary.

    Handles missing timestamps gracefully and computes duration
    when both start and complete times are available.

    Args:
        trial: Frozen trial from study

    Returns:
        Dictionary with trial information
    """
    duration = None
    if trial.datetime_complete and trial.datetime_start:
        duration = (trial.datetime_complete - trial.datetime_start).total_seconds()

    return {
        "number": trial.number,
        "value": trial.value,
        "params": trial.params,
        "state": trial.state.name,
        "datetime_start": trial.datetime_start.isoformat() if trial.datetime_start else None,
        "datetime_complete": (
            trial.datetime_complete.isoformat() if trial.datetime_complete else None
        ),
        "duration_seconds": duration,
    }


def build_top_trials_dataframe(
    sorted_trials: list[optuna.trial.FrozenTrial], metric_name: str
) -> pd.DataFrame:
    """
    Build DataFrame from sorted trials.

    Args:
        sorted_trials: list of trials sorted by performance
        metric_name: Name of optimization metric (for column header)

    Returns:
        Pandas DataFrame with trial comparison data
    """
    rows = []
    for rank, trial in enumerate(sorted_trials, 1):
        row = {
            "Rank": rank,
            "Trial": trial.number,
            f"{metric_name.upper()}": trial.value,
        }
        row.update(trial.params)

        # Add duration if available
        if trial.datetime_complete and trial.datetime_start:
            duration = (trial.datetime_complete - trial.datetime_start).total_seconds()
            row["Duration (s)"] = int(duration)

        rows.append(row)

    return pd.DataFrame(rows)
