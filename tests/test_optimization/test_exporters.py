"""
Test Suite for Exporting Configurations and Study Information.

This module contains test cases for verifying the export functionality of various
configurations such as best config, study summary, and top trials. It ensures
correct behavior of configuration serialization and export to formats such as YAML, JSON, and Excel.
"""

from __future__ import annotations

import json
from datetime import datetime
from unittest.mock import MagicMock, patch

import optuna
import pandas as pd
import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from orchard.core import Config, RunPaths
from orchard.optimization import (
    export_best_config,
    export_study_summary,
    export_top_trials,
)
from orchard.optimization.orchestrator.exporters import (
    TrialData,
    _auto_adjust_column_widths,
    _write_styled_rows,
    build_best_config_dict,
    build_best_trial_data,
    build_top_trials_dataframe,
)


# FIXTURES
@pytest.fixture
def study():
    """Fixture for creating a mock Optuna Study object."""
    study = MagicMock(spec=optuna.Study)
    trial_mock = MagicMock(spec=optuna.trial.FrozenTrial)
    trial_mock.state = optuna.trial.TrialState.COMPLETE
    trial_mock.params = {"param1": 0.1, "param2": 0.2}
    trial_mock.value = 0.8
    trial_mock.number = 1
    trial_mock.datetime_start = None
    trial_mock.datetime_complete = None

    study.best_trial = trial_mock
    study.best_params = trial_mock.params
    study.trials = [trial_mock]
    study.study_name = "test_study"
    study.direction = optuna.study.StudyDirection.MAXIMIZE
    return study


@pytest.fixture
def paths(tmpdir):
    """Fixture for creating a mock RunPaths object."""
    paths = MagicMock(spec=RunPaths)
    paths.reports = tmpdir.mkdir("reports")
    return paths


@pytest.fixture
def config():
    """Fixture for creating a valid Config object with ModelConfig and TrainingConfig."""
    model_config = {
        "name": "resnet_18",
        "pretrained": True,
        "dropout": 0.2,
        "weight_variant": None,
    }

    training_config = {"epochs": 10, "mixup_epochs": 5}

    return Config(model=model_config, training=training_config)


# TESTS
@pytest.mark.unit
def test_export_study_summary(study, paths):
    """Test export of study summary to JSON."""
    export_study_summary(study, paths)

    output_path = paths.reports / "study_summary.json"
    assert output_path.exists()

    with open(output_path, "r") as f:
        summary = json.load(f)
        assert "study_name" in summary
        assert "direction" in summary
        assert "trials" in summary
        assert len(summary["trials"]) == 1
        assert "best_trial" in summary


@pytest.mark.unit
def test_export_top_trials(study, paths):
    """Test export of top trials to Excel."""
    export_top_trials(study, paths, metric_name="accuracy", top_k=1)

    output_path = paths.reports / "top_10_trials.xlsx"
    assert output_path.exists()

    df = pd.read_excel(output_path)
    assert "Rank" in df.columns
    assert "Trial" in df.columns
    assert "ACCURACY" in df.columns


@pytest.mark.unit
def test_export_best_config_invalid_config(study, paths):
    """Test handling of invalid config during export."""

    invalid_model_config = "invalid_model"

    invalid_training_config = {"epochs": -10}

    with pytest.raises(ValidationError):
        invalid_config = Config(model=invalid_model_config, training=invalid_training_config)
        export_best_config(study, invalid_config, paths)

    output_path = paths.reports / "best_config.yaml"
    assert not output_path.exists()


@pytest.mark.unit
def test_export_study_summary_no_completed_trials(study, paths):
    """Test export when no completed trials exist."""
    study.trials = []
    export_study_summary(study, paths)

    output_path = paths.reports / "study_summary.json"
    assert output_path.exists()

    with open(output_path, "r") as f:
        summary = json.load(f)
        assert "n_completed" in summary
        assert summary["n_completed"] == 0
        assert "trials" in summary
        assert len(summary["trials"]) == 0


@pytest.mark.unit
def test_export_top_trials_no_completed_trials(study, paths):
    """Test export when no completed trials exist."""
    study.trials = []
    export_top_trials(study, paths, metric_name="accuracy", top_k=1)

    output_path = paths.reports / "top_10_trials.xlsx"
    assert not output_path.exists()


@pytest.mark.unit
def test_build_best_trial_data_value_error():
    """Test build_best_trial_data handles ValueError from study.best_trial."""
    study = MagicMock(spec=optuna.Study)
    study.best_trial.side_effect = ValueError("No trials")

    completed = []

    result = build_best_trial_data(study, completed)

    assert result is None


@pytest.mark.unit
def test_build_best_trial_data_value_error_with_completed_trials():
    """Test build_best_trial_data handles ValueError even with completed trials."""
    study = MagicMock(spec=optuna.Study)

    trial_mock = MagicMock(spec=optuna.trial.FrozenTrial)
    trial_mock.state = optuna.trial.TrialState.COMPLETE
    completed = [trial_mock]

    from unittest.mock import PropertyMock

    type(study).best_trial = PropertyMock(side_effect=ValueError("Corrupted study"))

    result = build_best_trial_data(study, completed)

    assert result is None


@pytest.mark.unit
def test_build_best_trial_data_value_error_direct_call():
    """Test build_best_trial_data ValueError exception path directly."""

    class BrokenStudy:
        @property
        def best_trial(self):
            raise ValueError("No best trial available")

    study = BrokenStudy()

    trial_mock = MagicMock(spec=optuna.trial.FrozenTrial)
    trial_mock.state = optuna.trial.TrialState.COMPLETE
    completed = [trial_mock]

    result = build_best_trial_data(study, completed)

    assert result is None


@pytest.mark.unit
def test_trial_data_from_trial_without_timestamps():
    """Test TrialData.from_trial when datetime fields are None."""
    trial = MagicMock(spec=optuna.trial.FrozenTrial)
    trial.number = 5
    trial.value = 0.95
    trial.params = {"lr": 0.001}
    trial.state = optuna.trial.TrialState.COMPLETE
    trial.datetime_start = None
    trial.datetime_complete = None

    result = TrialData.from_trial(trial)

    assert result.number == 5
    assert result.value == pytest.approx(0.95)
    assert result.params == {"lr": 0.001}
    assert result.state == "COMPLETE"
    assert result.datetime_start is None
    assert result.datetime_complete is None
    assert result.duration_seconds is None


@pytest.mark.unit
def test_trial_data_from_trial_with_only_start_time():
    """Test TrialData.from_trial when only start time is available."""
    trial = MagicMock(spec=optuna.trial.FrozenTrial)
    trial.number = 5
    trial.value = 0.95
    trial.params = {"lr": 0.001}
    trial.state = optuna.trial.TrialState.RUNNING
    trial.datetime_start = datetime(2024, 1, 1, 12, 0, 0)
    trial.datetime_complete = None

    result = TrialData.from_trial(trial)

    assert result.datetime_start is not None
    assert result.datetime_complete is None
    assert result.duration_seconds is None


@pytest.mark.unit
def test_build_top_trials_dataframe_without_duration():
    """Test build_top_trials_dataframe when trials have no timestamps."""
    trial1 = MagicMock(spec=optuna.trial.FrozenTrial)
    trial1.number = 1
    trial1.value = 0.95
    trial1.params = {"lr": 0.001, "dropout": 0.2}
    trial1.datetime_start = None
    trial1.datetime_complete = None

    trial2 = MagicMock(spec=optuna.trial.FrozenTrial)
    trial2.number = 2
    trial2.value = 0.93
    trial2.params = {"lr": 0.002, "dropout": 0.3}
    trial2.datetime_start = None
    trial2.datetime_complete = None

    sorted_trials = [trial1, trial2]

    df = build_top_trials_dataframe(sorted_trials, "auc")

    assert len(df) == 2
    assert "Rank" in df.columns
    assert "Trial" in df.columns
    assert "AUC" in df.columns
    assert "Duration (s)" not in df.columns


@pytest.mark.unit
def test_build_top_trials_dataframe_with_mixed_durations():
    """Test build_top_trials_dataframe with some trials having duration, some not."""
    trial1 = MagicMock(spec=optuna.trial.FrozenTrial)
    trial1.number = 1
    trial1.value = 0.95
    trial1.params = {"lr": 0.001}
    trial1.datetime_start = datetime(2024, 1, 1, 12, 0, 0)
    trial1.datetime_complete = datetime(2024, 1, 1, 12, 5, 30)
    trial2 = MagicMock(spec=optuna.trial.FrozenTrial)
    trial2.number = 2
    trial2.value = 0.93
    trial2.params = {"lr": 0.002}
    trial2.datetime_start = None
    trial2.datetime_complete = None

    sorted_trials = [trial1, trial2]

    df = build_top_trials_dataframe(sorted_trials, "accuracy")

    assert len(df) == 2
    assert "Duration (s)" in df.columns
    assert df.loc[0, "Duration (s)"] == 330
    assert pd.isna(df.loc[1, "Duration (s)"]) or "Duration (s)" not in df.iloc[1]


@pytest.mark.unit
def test_export_best_config_no_completed_trials_integration(minimal_config, paths):
    """Test export_best_config returns None when no completed trials exist."""
    study = MagicMock()

    trial1 = MagicMock()
    trial1.state = optuna.trial.TrialState.FAIL

    trial2 = MagicMock()
    trial2.state = optuna.trial.TrialState.PRUNED

    study.trials = [trial1, trial2]

    with patch("orchard.optimization.orchestrator.exporters.logger") as mock_logger:
        result = export_best_config(study, minimal_config, paths)
        assert result is None

        mock_logger.warning.assert_called_once()


@pytest.mark.unit
def test_export_best_config_success_path(minimal_config, paths, tmp_path):
    """Test export_best_config creates YAML when trials exist."""
    study = MagicMock()
    study.best_params = {
        "learning_rate": 0.001,
        "batch_size": 32,
        "weight_decay": 0.0001,
    }

    trial = MagicMock()
    trial.state = optuna.trial.TrialState.COMPLETE
    trial.value = 0.95
    study.trials = [trial]

    paths.reports = tmp_path / "reports"
    paths.reports.mkdir(parents=True, exist_ok=True)

    with patch("orchard.optimization.orchestrator.exporters.build_best_config_dict") as mock_build:
        with patch("orchard.optimization.orchestrator.exporters.save_config_as_yaml") as mock_save:
            mock_build.return_value = {
                "training": {"learning_rate": 0.001, "batch_size": 32},
                "dataset": {"name": "test"},
                "architecture": {"name": "resnet"},
            }

            result = export_best_config(study, minimal_config, paths)

            assert result == paths.reports / "best_config.yaml"

            mock_build.assert_called_once_with(study.best_params, minimal_config)
            mock_save.assert_called_once()


@pytest.mark.unit
def test_export_top_trials_all_type_branches(paths, tmp_path):
    """Test that all formatting branches (float, int, bool, string) are covered."""
    study = MagicMock(spec=optuna.Study)
    study.direction = optuna.study.StudyDirection.MAXIMIZE

    trial = MagicMock(spec=optuna.trial.FrozenTrial)
    trial.state = optuna.trial.TrialState.COMPLETE
    trial.number = 1
    trial.value = 0.9567
    trial.params = {
        "learning_rate": 0.001234,
        "batch_size": 32,
        "dropout": 0.25,
        "use_amp": True,
        "optimizer": "adamw",
    }
    trial.datetime_start = datetime(2024, 1, 1, 12, 0, 0)
    trial.datetime_complete = datetime(2024, 1, 1, 12, 5, 30)

    study.trials = [trial]

    paths.reports = tmp_path / "reports"
    paths.reports.mkdir(parents=True, exist_ok=True)

    export_top_trials(study, paths, metric_name="auc", top_k=1)

    output_path = paths.reports / "top_10_trials.xlsx"
    assert output_path.exists()

    df = pd.read_excel(output_path)
    assert len(df) == 1

    assert df.loc[0, "Rank"] == 1
    assert df.loc[0, "Trial"] == 1
    assert df.loc[0, "AUC"] == pytest.approx(0.9567)
    assert df.loc[0, "learning_rate"] == pytest.approx(0.001234)
    assert df.loc[0, "batch_size"] == 32
    assert df.loc[0, "Duration (s)"] == 330


@pytest.mark.unit
def test_build_best_config_dict():
    """Test building config dict from best trial params."""
    mock_cfg = MagicMock()
    mock_cfg.training.epochs = 50
    mock_cfg.model_dump = MagicMock(
        return_value={
            "training": {"epochs": 50},
            "architecture": {},
            "augmentation": {},
        }
    )

    params = {"learning_rate": 0.001, "dropout": 0.5}
    config_dict = build_best_config_dict(params, mock_cfg)

    assert config_dict["training"]["learning_rate"] == pytest.approx(0.001)
    assert config_dict["architecture"]["dropout"] == pytest.approx(0.5)
    assert config_dict["training"]["epochs"] == 50


@pytest.mark.unit
def test_build_best_config_dict_epochs_key():
    """Test that build_best_config_dict writes to the exact 'epochs' key."""
    mock_cfg = MagicMock()
    mock_cfg.training.epochs = 100
    mock_cfg.model_dump = MagicMock(
        return_value={
            "training": {"epochs": 5, "learning_rate": 0.01},
            "architecture": {},
        }
    )

    params = {"learning_rate": 0.002}
    result = build_best_config_dict(params, mock_cfg)

    assert "epochs" in result["training"]
    assert result["training"]["epochs"] == 100
    assert "EPOCHS" not in result["training"]
    assert "XXepochsXX" not in result["training"]
    assert isinstance(result, dict)


@pytest.mark.unit
def test_export_study_summary_json_structure(study, tmp_path):
    """Test that export_study_summary produces correct JSON keys and formatting."""
    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    export_study_summary(study, paths)

    output_path = paths.reports / "study_summary.json"
    raw = output_path.read_text()

    summary = json.loads(raw)
    assert summary["study_name"] == "test_study"
    assert summary["direction"] == "MAXIMIZE"
    assert "n_trials" in summary
    assert summary["n_trials"] == 1
    assert summary["n_completed"] == 1
    assert summary["best_trial"] is not None
    assert summary["best_trial"]["value"] == 0.8

    # Verify indent=2 (not None, not 3)
    lines = raw.split("\n")
    indented_lines = [
        line for line in lines if line.startswith("  ") and not line.startswith("    ")
    ]
    assert len(indented_lines) > 0, "JSON should be indented with 2 spaces"
    assert not any(
        line.startswith("   ") and not line.startswith("    ") for line in lines
    ), "No 3-space indentation expected"


@pytest.mark.unit
def test_export_study_summary_best_trial_data(tmp_path):
    """Test that best_trial_data is correctly computed from study, not set to None."""
    study = MagicMock(spec=optuna.Study)
    trial = MagicMock(spec=optuna.trial.FrozenTrial)
    trial.state = optuna.trial.TrialState.COMPLETE
    trial.params = {"lr": 0.01}
    trial.value = 0.9
    trial.number = 0
    trial.datetime_start = datetime(2024, 1, 1, 10, 0, 0)
    trial.datetime_complete = datetime(2024, 1, 1, 10, 5, 0)
    study.best_trial = trial
    study.trials = [trial]
    study.study_name = "test"
    study.direction = optuna.study.StudyDirection.MAXIMIZE

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    export_study_summary(study, paths)

    output_path = paths.reports / "study_summary.json"
    summary = json.loads(output_path.read_text())

    assert summary["best_trial"] is not None
    assert summary["best_trial"]["number"] == 0
    assert summary["best_trial"]["value"] == 0.9
    assert summary["best_trial"]["datetime_start"] is not None
    assert summary["best_trial"]["datetime_complete"] is not None
    assert summary["best_trial"]["duration_seconds"] == 300.0


@pytest.mark.unit
def test_export_top_trials_sorting_minimize(tmp_path):
    """Test that MINIMIZE direction sorts trials ascending."""
    study = MagicMock(spec=optuna.Study)
    study.direction = optuna.study.StudyDirection.MINIMIZE

    trial1 = MagicMock(spec=optuna.trial.FrozenTrial)
    trial1.state = optuna.trial.TrialState.COMPLETE
    trial1.number = 1
    trial1.value = 0.5
    trial1.params = {"lr": 0.1}
    trial1.datetime_start = None
    trial1.datetime_complete = None

    trial2 = MagicMock(spec=optuna.trial.FrozenTrial)
    trial2.state = optuna.trial.TrialState.COMPLETE
    trial2.number = 2
    trial2.value = 0.3
    trial2.params = {"lr": 0.01}
    trial2.datetime_start = None
    trial2.datetime_complete = None

    study.trials = [trial1, trial2]

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    export_top_trials(study, paths, metric_name="loss", top_k=10)

    output_path = paths.reports / "top_10_trials.xlsx"
    df = pd.read_excel(output_path)
    assert df.loc[0, "LOSS"] == pytest.approx(0.3)
    assert df.loc[1, "LOSS"] == pytest.approx(0.5)


@pytest.mark.unit
def test_export_top_trials_filters_nan_and_none(tmp_path):
    """Test that trials with NaN or None values are excluded from top trials."""
    study = MagicMock(spec=optuna.Study)
    study.direction = optuna.study.StudyDirection.MAXIMIZE

    good_trial = MagicMock(spec=optuna.trial.FrozenTrial)
    good_trial.state = optuna.trial.TrialState.COMPLETE
    good_trial.number = 1
    good_trial.value = 0.9
    good_trial.params = {"lr": 0.01}
    good_trial.datetime_start = None
    good_trial.datetime_complete = None

    nan_trial = MagicMock(spec=optuna.trial.FrozenTrial)
    nan_trial.state = optuna.trial.TrialState.COMPLETE
    nan_trial.number = 2
    nan_trial.value = float("nan")
    nan_trial.params = {"lr": 0.05}
    nan_trial.datetime_start = None
    nan_trial.datetime_complete = None

    none_trial = MagicMock(spec=optuna.trial.FrozenTrial)
    none_trial.state = optuna.trial.TrialState.COMPLETE
    none_trial.number = 3
    none_trial.value = None
    none_trial.params = {"lr": 0.1}
    none_trial.datetime_start = None
    none_trial.datetime_complete = None

    study.trials = [good_trial, nan_trial, none_trial]

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    export_top_trials(study, paths, metric_name="acc", top_k=10)

    output_path = paths.reports / "top_10_trials.xlsx"
    df = pd.read_excel(output_path)
    assert len(df) == 1
    assert df.loc[0, "ACC"] == pytest.approx(0.9)


@pytest.mark.unit
def test_export_top_trials_top_k_slicing(tmp_path):
    """Test that top_k correctly limits the number of exported trials."""
    study = MagicMock(spec=optuna.Study)
    study.direction = optuna.study.StudyDirection.MAXIMIZE

    trials = []
    for i in range(5):
        t = MagicMock(spec=optuna.trial.FrozenTrial)
        t.state = optuna.trial.TrialState.COMPLETE
        t.number = i
        t.value = float(i) / 10
        t.params = {"lr": 0.01}
        t.datetime_start = None
        t.datetime_complete = None
        trials.append(t)

    study.trials = trials

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    export_top_trials(study, paths, metric_name="acc", top_k=2)

    df = pd.read_excel(paths.reports / "top_10_trials.xlsx")
    assert len(df) == 2


@pytest.mark.unit
def test_write_styled_rows_header_formatting(tmp_path):
    """Test that header row has correct fill, font, and alignment."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    df = pd.DataFrame({"Col1": [1.5, 2], "Col2": ["a", "b"]})
    _write_styled_rows(ws, df)

    header_cell = ws.cell(row=1, column=1)
    assert header_cell.fill.start_color.rgb == "00D7E4BC"
    assert header_cell.font.bold is True
    assert header_cell.alignment.horizontal == "center"
    assert header_cell.alignment.vertical == "center"
    assert header_cell.border.left.style == "thin"
    assert header_cell.border.right.style == "thin"
    assert header_cell.border.top.style == "thin"
    assert header_cell.border.bottom.style == "thin"


@pytest.mark.unit
def test_write_styled_rows_body_formatting(tmp_path):
    """Test that body rows have correct alignment and number formatting."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    df = pd.DataFrame({"FloatCol": [3.14], "IntCol": [42], "BoolCol": [True]})
    _write_styled_rows(ws, df)

    # Body row is row 2
    float_cell = ws.cell(row=2, column=1)
    assert float_cell.alignment.horizontal == "left"
    assert float_cell.alignment.vertical == "center"
    assert float_cell.alignment.wrap_text is True
    assert float_cell.number_format == "0.0000"
    assert float_cell.border.left.style == "thin"

    int_cell = ws.cell(row=2, column=2)
    assert int_cell.number_format == "0"

    # Bool should NOT get int formatting (bool is subclass of int)
    bool_cell = ws.cell(row=2, column=3)
    assert bool_cell.number_format != "0"


@pytest.mark.unit
def test_write_styled_rows_cell_positions(tmp_path):
    """Test that cells are written starting at column 1 (not 2)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    df = pd.DataFrame({"A": [10], "B": [20]})
    _write_styled_rows(ws, df)

    assert ws.cell(row=1, column=1).value == "A"
    assert ws.cell(row=1, column=2).value == "B"
    assert ws.cell(row=2, column=1).value == 10
    assert ws.cell(row=2, column=2).value == 20


@pytest.mark.unit
def test_write_styled_rows_dataframe_no_index(tmp_path):
    """Test that index=False is respected (no index column in output)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    df = pd.DataFrame({"X": [1, 2, 3]})
    _write_styled_rows(ws, df)

    # Should have 1 header row + 3 data rows = 4 rows, 1 column
    assert ws.cell(row=1, column=1).value == "X"
    assert ws.cell(row=4, column=1).value == 3
    # Column 2 should be empty for all rows
    assert ws.cell(row=1, column=2).value is None


@pytest.mark.unit
def test_auto_adjust_column_widths_values():
    """Test that column widths are computed correctly with +2, min 12, max 50."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    # Short content → should be min 12
    ws.cell(row=1, column=1, value="Hi")
    # Medium content → length + 2
    ws.cell(row=1, column=2, value="A" * 20)
    # Very long content → should be capped at 50
    ws.cell(row=1, column=3, value="B" * 100)
    # Empty cell → should be min 12
    ws.cell(row=1, column=4, value=None)

    _auto_adjust_column_widths(ws)

    assert ws.column_dimensions["A"].width == 12  # max(2+2, 12) = 12
    assert ws.column_dimensions["B"].width == 22  # max(20+2, 12) = 22
    assert ws.column_dimensions["C"].width == 50  # min(102, 50) = 50
    assert ws.column_dimensions["D"].width == 12  # empty → 12


@pytest.mark.unit
def test_auto_adjust_column_widths_first_cell_letter():
    """Test that column letter is taken from first cell (index 0), not index 1."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Header")
    ws.cell(row=2, column=1, value="Data")

    _auto_adjust_column_widths(ws)

    # Column A width should be based on max content length
    assert ws.column_dimensions["A"].width == max(len("Header") + 2, 12)


@pytest.mark.unit
def test_export_top_trials_worksheet_title(tmp_path):
    """Test that the worksheet title is exactly 'Top Trials'."""
    study = MagicMock(spec=optuna.Study)
    study.direction = optuna.study.StudyDirection.MAXIMIZE

    trial = MagicMock(spec=optuna.trial.FrozenTrial)
    trial.state = optuna.trial.TrialState.COMPLETE
    trial.number = 1
    trial.value = 0.9
    trial.params = {"lr": 0.01}
    trial.datetime_start = None
    trial.datetime_complete = None
    study.trials = [trial]

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    export_top_trials(study, paths, metric_name="acc", top_k=1)

    wb = load_workbook(paths.reports / "top_10_trials.xlsx")
    assert wb.active.title == "Top Trials"
    wb.close()


@pytest.mark.unit
def test_export_best_config_calls_save_with_correct_args(tmp_path):
    """Test export_best_config passes correct Config and path to save_config_as_yaml."""
    study = MagicMock()
    study.best_params = {"learning_rate": 0.001}

    trial = MagicMock()
    trial.state = optuna.trial.TrialState.COMPLETE
    trial.value = 0.9
    study.trials = [trial]

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    cfg = MagicMock()
    cfg.training.epochs = 50
    cfg.model_dump.return_value = {
        "training": {"epochs": 5, "learning_rate": 0.01},
    }

    with patch("orchard.optimization.orchestrator.exporters.Config") as mock_config_cls:
        mock_config_instance = MagicMock()
        mock_config_cls.return_value = mock_config_instance

        with patch("orchard.optimization.orchestrator.exporters.save_config_as_yaml") as mock_save:
            result = export_best_config(study, cfg, paths)

            expected_path = paths.reports / "best_config.yaml"
            assert result == expected_path

            # Verify Config was constructed (not None)
            mock_config_cls.assert_called_once()
            # Verify save was called with the Config instance AND the correct path
            mock_save.assert_called_once_with(mock_config_instance, expected_path)


@pytest.mark.unit
def test_export_top_trials_excel_formatting_roundtrip(tmp_path):
    """Test full Excel roundtrip: verify formatting via openpyxl read-back."""
    study = MagicMock(spec=optuna.Study)
    study.direction = optuna.study.StudyDirection.MAXIMIZE

    trial = MagicMock(spec=optuna.trial.FrozenTrial)
    trial.state = optuna.trial.TrialState.COMPLETE
    trial.number = 1
    trial.value = 0.95
    trial.params = {"lr": 0.001, "batch_size": 32}
    trial.datetime_start = datetime(2024, 1, 1, 12, 0, 0)
    trial.datetime_complete = datetime(2024, 1, 1, 12, 5, 0)
    study.trials = [trial]

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    export_top_trials(study, paths, metric_name="auc", top_k=1)

    wb = load_workbook(paths.reports / "top_10_trials.xlsx")
    ws = wb.active

    # Header row formatting
    h1 = ws.cell(row=1, column=1)
    assert h1.fill.start_color.rgb == "00D7E4BC"
    assert h1.fill.fill_type == "solid"
    assert h1.font.bold is True
    assert h1.alignment.horizontal == "center"

    # Body row formatting
    b_float = ws.cell(row=2, column=3)  # AUC column (float)
    assert b_float.alignment.horizontal == "left"
    assert b_float.number_format == "0.0000"

    # Border on all cells
    for row in ws.iter_rows(min_row=1, max_row=2, max_col=6):
        for cell in row:
            if cell.value is not None:
                assert cell.border.left.style == "thin"
                assert cell.border.right.style == "thin"
                assert cell.border.top.style == "thin"
                assert cell.border.bottom.style == "thin"

    wb.close()


@pytest.mark.unit
def test_build_top_trials_dataframe_only_complete_set():
    """Test that duration is only added when both start AND complete are set."""
    trial = MagicMock(spec=optuna.trial.FrozenTrial)
    trial.number = 1
    trial.value = 0.9
    trial.params = {"lr": 0.01}
    trial.datetime_start = datetime(2024, 1, 1, 12, 0, 0)
    trial.datetime_complete = None  # only start, no complete

    df = build_top_trials_dataframe([trial], "acc")
    assert "Duration (s)" not in df.columns


@pytest.mark.unit
def test_write_styled_rows_header_fill_end_color():
    """Test that header fill end_color is set to the exact expected value."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    df = pd.DataFrame({"A": [1]})
    _write_styled_rows(ws, df)

    header = ws.cell(row=1, column=1)
    # end_color must be set (not None/default) and match start_color
    assert header.fill.end_color.rgb == "00D7E4BC"


@pytest.mark.unit
def test_export_top_trials_maximize_descending_order(tmp_path):
    """Test that MAXIMIZE direction sorts trials in descending order."""
    study = MagicMock(spec=optuna.Study)
    study.direction = optuna.study.StudyDirection.MAXIMIZE

    trial_low = MagicMock(spec=optuna.trial.FrozenTrial)
    trial_low.state = optuna.trial.TrialState.COMPLETE
    trial_low.number = 1
    trial_low.value = 0.3
    trial_low.params = {"lr": 0.1}
    trial_low.datetime_start = None
    trial_low.datetime_complete = None

    trial_high = MagicMock(spec=optuna.trial.FrozenTrial)
    trial_high.state = optuna.trial.TrialState.COMPLETE
    trial_high.number = 2
    trial_high.value = 0.9
    trial_high.params = {"lr": 0.01}
    trial_high.datetime_start = None
    trial_high.datetime_complete = None

    study.trials = [trial_low, trial_high]

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    export_top_trials(study, paths, metric_name="acc", top_k=10)

    df = pd.read_excel(paths.reports / "top_10_trials.xlsx")
    # MAXIMIZE: highest value first
    assert df.loc[0, "ACC"] == pytest.approx(0.9)
    assert df.loc[1, "ACC"] == pytest.approx(0.3)


@pytest.mark.unit
def test_export_top_trials_default_top_k(tmp_path):
    """Test that the default top_k is exactly 10."""
    study = MagicMock(spec=optuna.Study)
    study.direction = optuna.study.StudyDirection.MAXIMIZE

    trials = []
    for i in range(15):
        t = MagicMock(spec=optuna.trial.FrozenTrial)
        t.state = optuna.trial.TrialState.COMPLETE
        t.number = i
        t.value = float(i)
        t.params = {"lr": 0.01}
        t.datetime_start = None
        t.datetime_complete = None
        trials.append(t)

    study.trials = trials

    paths = MagicMock(spec=RunPaths)
    paths.reports = tmp_path / "reports"
    paths.reports.mkdir()

    # Call without explicit top_k — should default to 10
    export_top_trials(study, paths, metric_name="acc")

    df = pd.read_excel(paths.reports / "top_10_trials.xlsx")
    assert len(df) == 10


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
